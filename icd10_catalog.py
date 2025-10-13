"""ICD/Operation catalog helpers for Registry Patient Connect."""
from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
from functools import lru_cache
from pathlib import Path
from typing import Dict, List, Sequence, Set, Tuple

# ---------------------------------------------------------------------------
# Core catalog data
# ---------------------------------------------------------------------------

_SPECIALTY_OPERATIONS: Dict[str, List[str]] = {
    "Surgery": [
        "Appendectomy (Laparoscopic)",
        "Cholecystectomy (Laparoscopic)",
        "Hemorrhoidectomy",
        "Breast Lumpectomy",
    ],
    "Orthopedics": [
        "Open Reduction Internal Fixation (ORIF) — Radius",
        "Total Knee Arthroplasty",
        "Arthroscopic Meniscectomy",
        "External Fixator Adjustment",
    ],
    "Urology": [
        "Transurethral Resection of Prostate (TURP)",
        "Cystolitholapaxy",
        "Percutaneous Nephrolithotomy",
    ],
    "OBGYN": [
        "Cesarean Section",
        "Total Abdominal Hysterectomy",
        "Diagnostic Laparoscopy",
    ],
    "ENT": [
        "Tonsillectomy",
        "Functional Endoscopic Sinus Surgery (FESS)",
        "Myringotomy with Tube Insertion",
    ],
    "Ophthalmology": [
        "Phacoemulsification with IOL",
        "Pterygium Excision",
        "Vitrectomy",
    ],
    "Maxillofacial": [
        "Open Reduction Internal Fixation — Mandible",
        "Le Fort I Osteotomy",
        "Temporomandibular Joint Arthroplasty",
    ],
}

_SPECIALTY_DIAGNOSES: Dict[str, List[str]] = {
    "Surgery": [
        "K35.80 - Acute appendicitis, unspecified",
        "K80.20 - Calculus of gallbladder without cholecystitis",
        "I84.90 - Hemorrhoids, unspecified",
    ],
    "Orthopedics": [
        "S52.50 - Fracture of distal radius",
        "M17.10 - Osteoarthritis of knee, unspecified",
        "S83.20 - Tear of meniscus, unspecified",
    ],
    "Urology": [
        "N40.1 - Benign prostatic hyperplasia with LUTS",
        "N20.0 - Calculus of kidney",
    ],
    "OBGYN": [
        "O82.0 - Cesarean delivery, unspecified",
        "N80.9 - Endometriosis, unspecified",
    ],
    "ENT": [
        "J35.01 - Chronic tonsillitis",
        "J32.9 - Chronic sinusitis, unspecified",
    ],
    "Ophthalmology": [
        "H25.9 - Senile cataract, unspecified",
        "H11.0 - Pterygium",
    ],
    "Maxillofacial": [
        "S02.609 - Fracture of mandible",
        "M26.609 - Temporomandibular joint disorder",
    ],
}

_OPERATION_TO_DIAGNOSES: Dict[str, List[str]] = {
    "Appendectomy (Laparoscopic)": ["K35.80 - Acute appendicitis, unspecified"],
    "Cholecystectomy (Laparoscopic)": ["K80.20 - Calculus of gallbladder without cholecystitis"],
    "Hemorrhoidectomy": ["I84.90 - Hemorrhoids, unspecified"],
    "Breast Lumpectomy": ["D05.9 - Carcinoma in situ of breast"],
    "Open Reduction Internal Fixation (ORIF) — Radius": ["S52.50 - Fracture of distal radius"],
    "Total Knee Arthroplasty": ["M17.10 - Osteoarthritis of knee, unspecified"],
    "Arthroscopic Meniscectomy": ["S83.20 - Tear of meniscus, unspecified"],
    "External Fixator Adjustment": ["S82.209 - Fracture of tibia"],
    "Transurethral Resection of Prostate (TURP)": ["N40.1 - Benign prostatic hyperplasia with LUTS"],
    "Cystolitholapaxy": ["N21.0 - Calculus in bladder"],
    "Percutaneous Nephrolithotomy": ["N20.0 - Calculus of kidney"],
    "Cesarean Section": ["O82.0 - Cesarean delivery, unspecified"],
    "Total Abdominal Hysterectomy": ["N85.2 - Hypertrophy of uterus"],
    "Diagnostic Laparoscopy": ["N80.9 - Endometriosis, unspecified"],
    "Tonsillectomy": ["J35.01 - Chronic tonsillitis"],
    "Functional Endoscopic Sinus Surgery (FESS)": ["J32.9 - Chronic sinusitis, unspecified"],
    "Myringotomy with Tube Insertion": ["H66.90 - Otitis media, unspecified"],
    "Phacoemulsification with IOL": ["H25.9 - Senile cataract, unspecified"],
    "Pterygium Excision": ["H11.0 - Pterygium"],
    "Vitrectomy": ["H43.1 - Vitreous hemorrhage"],
    "Open Reduction Internal Fixation — Mandible": ["S02.609 - Fracture of mandible"],
    "Le Fort I Osteotomy": ["M26.212 - Maxillary hyperplasia"],
    "Temporomandibular Joint Arthroplasty": ["M26.609 - Temporomandibular joint disorder"],
}


def _resolve_portable_path(env_key: str, default_rel: str = "") -> Path:
    """Resolve ``env_key`` to a portable :class:`Path`."""

    env_value = os.getenv(env_key, "").strip()
    candidates: List[Path] = []

    if env_value:
        expanded = Path(os.path.expandvars(os.path.expanduser(env_value)))
        candidates.append(expanded)

    app_dir = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    if env_value and not Path(env_value).is_absolute():
        candidates.append(app_dir / env_value)

    if default_rel:
        candidates.append(app_dir / default_rel)

    home = Path.home()
    candidates.append(home / "Desktop" / "my-project" / "ICD9_ICD10_by_Specialty.xlsx")

    if env_value:
        candidates.append(Path(env_value))

    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate

    return candidates[0] if candidates else app_dir


def _xlsx_rows(path: Path, sheet: str) -> List[List[str]]:
    """Read ``sheet`` rows from ``path`` using :mod:`openpyxl`."""

    try:  # pragma: no cover - optional dependency
        import openpyxl  # type: ignore
    except Exception:
        return []

    if not path.exists():
        return []

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet]
        result: List[List[str]] = []
        for row in worksheet.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            result.append([("" if cell is None else str(cell)).strip() for cell in row])
        return result
    except Exception:
        return []


def _user_db_path() -> Path:
    configured = os.getenv("ICD_CUSTOM_DB_PATH", "").strip()
    if configured:
        return Path(os.path.expandvars(os.path.expanduser(configured)))

    if os.name == "nt":
        base = Path(os.getenv("APPDATA", Path.home() / "AppData" / "Roaming"))
    else:
        base = Path(os.getenv("XDG_DATA_HOME", Path.home() / ".local" / "share"))

    target = base / "SurgiBot" / "icd_user_additions.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def _load_user_db() -> Dict[str, Dict[str, List[str]]]:
    fp = _user_db_path()
    if not fp.exists():
        return {"diagnosis": {}, "operation": {}}
    try:
        return json.loads(fp.read_text(encoding="utf-8"))
    except Exception:
        return {"diagnosis": {}, "operation": {}}


def _atomic_json_write(fp: Path, payload: Dict[str, Dict[str, List[str]]]) -> None:
    fp.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=str(fp.parent)) as tmp:
        json.dump(payload, tmp, ensure_ascii=False, indent=2)
        tmp_path = Path(tmp.name)
    shutil.move(str(tmp_path), fp)


def add_custom_entry(kind: str, specialty: str, text: str) -> bool:
    """Persist ``text`` for ``kind``/``specialty``. Returns ``True`` if stored."""

    kind_key = (kind or "").strip().lower()
    if kind_key not in {"diagnosis", "operation"}:
        return False
    specialty_key = (specialty or "").strip()
    entry = (text or "").strip()
    if not specialty_key or not entry:
        return False

    db = _load_user_db()
    bucket = db.setdefault(kind_key, {})
    items = bucket.setdefault(specialty_key, [])
    if entry in items:
        return False
    items.append(entry)
    _atomic_json_write(_user_db_path(), db)
    return True


def get_custom_list(kind: str, specialty: str) -> List[str]:
    kind_key = (kind or "").strip().lower()
    specialty_key = (specialty or "").strip()
    if not specialty_key:
        return []
    db = _load_user_db()
    return list(db.get(kind_key, {}).get(specialty_key, []))


_SPECIALTY_SHEETS: Dict[str, Tuple[str, str]] = {
    "Surgery": ("Surgery_ICD10", "Surgery_ICD9"),
    "Orthopedics": ("Orthopedics_ICD10", "Orthopedics_ICD9"),
    "Urology": ("Urology_ICD10", "Urology_ICD9"),
    "ENT": ("ENT_ICD10", "ENT_ICD9"),
    "Obstetrics-Gynecology": ("OBGYN_ICD10", "OBGYN_ICD9"),
    "Ophthalmology": ("Ophthalmology_ICD10", "Ophthalmology_ICD9"),
    "Maxillofacial": ("Maxillofacial_ICD10", "Maxillofacial_ICD9"),
}


def _pack_list_from_sheet(rows: List[List[str]]) -> List[str]:
    if not rows:
        return []

    header = [cell.lower() for cell in rows[0]]

    def _col(names: List[str]) -> int:
        for name in names:
            if name in header:
                return header.index(name)
        return -1

    code_idx = _col(["code", "codeplain", "icd-10", "icd10", "icd9", "รหัส", "code "])
    term_idx = _col([
        "term",
        "long description (valid icd-9 fy2025)",
        "long description",
        "รายละเอียด",
        "name",
    ])

    start_row = 1 if code_idx != -1 or term_idx != -1 else 0
    packed: List[str] = []

    for row in rows[start_row:]:
        if not row:
            continue
        code = row[code_idx].strip() if 0 <= code_idx < len(row) else ""
        desc = row[term_idx].strip() if 0 <= term_idx < len(row) else ""
        if code and desc:
            packed.append(f"{code} - {desc}")
        elif desc:
            packed.append(desc)
        elif code:
            packed.append(code)

    unique: List[str] = []
    seen: Set[str] = set()
    for item in packed:
        if item and item not in seen:
            seen.add(item)
            unique.append(item)
    return unique


def load_specialty_catalog(specialty: str) -> Tuple[List[str], List[str]]:
    if os.getenv("USE_SPECIALTY_CATALOG", "0").strip() != "1":
        return [], []

    catalog_path = _resolve_portable_path(
        "SPECIALTY_CATALOG_PATH", default_rel="ICD9_ICD10_by_Specialty.xlsx"
    )
    sheets = _SPECIALTY_SHEETS.get(specialty)
    if not sheets or not catalog_path.exists():
        return [], []

    icd10_sheet, icd9_sheet = sheets
    icd10_rows = _xlsx_rows(catalog_path, icd10_sheet)
    icd9_rows = _xlsx_rows(catalog_path, icd9_sheet)
    return _pack_list_from_sheet(icd10_rows), _pack_list_from_sheet(icd9_rows)


def _normalize_specialty(key: str) -> str:
    return (key or "Surgery").strip().title()


@lru_cache(maxsize=None)
def all_operations() -> List[str]:
    ops: Set[str] = set()
    for values in _SPECIALTY_OPERATIONS.values():
        ops.update(values)
    return sorted(ops)


ALL_OPERATIONS: List[str] = all_operations()


def operation_suggestions(specialty_key: str) -> List[str]:
    key = _normalize_specialty(specialty_key)
    ops = _SPECIALTY_OPERATIONS.get(key)
    if ops:
        return list(ops)
    return list(ALL_OPERATIONS)


def diagnosis_suggestions(specialty_key: str, operations: Sequence[str] | None = None) -> List[str]:
    key = _normalize_specialty(specialty_key)
    suggestions: Set[str] = set(_SPECIALTY_DIAGNOSES.get(key, []))
    for op in operations or []:
        suggestions.update(_OPERATION_TO_DIAGNOSES.get(op, []))
    return sorted(suggestions)


# ---------------------------------------------------------------------------
# Excel loader for ICD-10-TM
# ---------------------------------------------------------------------------

def load_icd10tm_xlsx(xlsx_path: str) -> List[str]:
    """Load ICD-10-TM entries from the given Excel file.

    Returns a list of strings formatted as ``"CODE - Description"``.
    When :mod:`openpyxl` is missing or the file cannot be read, an empty
    list is returned so the caller can fall back to the in-memory catalog.
    """

    try:
        import openpyxl  # type: ignore
    except Exception:
        return []

    path = Path(xlsx_path)
    if not path.exists():
        return []

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        results: List[str] = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            values = [str(cell).strip() for cell in row if cell is not None]
            if not values:
                continue

            code = ""
            label = ""
            for value in values:
                if not code:
                    code = value
                    continue
                if not label:
                    label = value
                    break

            if not code or not label:
                continue

            lower_code = code.lower()
            lower_label = label.lower()
            if lower_code in {
                "code",
                "icd",
                "icd10",
                "icd-10",
                "icd10tm",
                "รหัส",
                "codeid",
                "diag",
                "diagnosis",
                "ชื่อโรค",
                "รายละเอียด",
            }:
                continue
            if lower_label in {
                "code",
                "icd",
                "icd10",
                "icd-10",
                "icd10tm",
                "รหัส",
                "codeid",
                "diag",
                "diagnosis",
                "ชื่อโรค",
                "รายละเอียด",
            }:
                continue

            results.append(f"{code} - {label}")

        return sorted({item for item in results if item}, key=str.lower)
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Excel loader for ICD-9 operations (valid + excluded)
# ---------------------------------------------------------------------------


def _read_xlsx_rows_icd9(xlsx_path: str):
    """Yield trimmed string rows from an Excel worksheet used for ICD-9 lists."""

    try:  # pragma: no cover - optional dependency
        import openpyxl  # type: ignore
    except Exception:
        return []

    path = Path(xlsx_path)
    if not path.exists():
        return []

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            yield [str(cell).strip() for cell in row if cell is not None]
    except Exception:
        return []


def load_icd9_ops(valid_path: str, exclude_path: str) -> List[str]:
    """Load ICD-9 operations from Excel workbooks.

    ``valid_path`` should point to the workbook that contains the canonical
    operation list, while ``exclude_path`` lists codes that must be removed.
    Returned values follow the ``"CODE - Description"`` format and are sorted
    by the normalized code for a consistent experience.
    """

    excluded: Set[str] = set()
    for values in _read_xlsx_rows_icd9(exclude_path):
        if not values:
            continue
        code = values[0].split()[0] if values else ""
        normalized = code.replace(".", "").strip().upper()
        if not normalized:
            continue
        if normalized in {"CODE", "ICD9", "ICD-9", "รหัส"}:
            continue
        excluded.add(normalized)

    items: Dict[str, str] = {}
    for values in _read_xlsx_rows_icd9(valid_path):
        if not values:
            continue
        raw_code = (values[0] if len(values) >= 1 else "").strip()
        raw_desc = (values[1] if len(values) >= 2 else "").strip()
        if not raw_code or not raw_desc:
            continue

        lowered = raw_code.lower()
        if lowered in {"code", "icd9", "icd-9", "operation", "รหัส", "หัตถการ"}:
            continue

        key = raw_code.replace(".", "").strip().upper()
        if key in excluded:
            continue

        items[key] = f"{raw_code} - {raw_desc}"

    return [items[idx] for idx in sorted(items.keys())]


def _merge_icd9_from_env_into_all_operations() -> None:
    """Append ICD-9 operations from environment-provided workbooks."""

    valid_path = os.getenv("ICD9_VALID_XLSX_PATH", "")
    exclude_path = os.getenv("ICD9_EXCLUDED_XLSX_PATH", "")
    if not valid_path or not exclude_path:
        return

    try:
        extra_ops = load_icd9_ops(valid_path, exclude_path)
    except Exception:
        extra_ops = []

    if not extra_ops:
        return

    merged: List[str] = list(dict.fromkeys(ALL_OPERATIONS + extra_ops))
    globals()["ALL_OPERATIONS"] = merged


_merge_icd9_from_env_into_all_operations()
